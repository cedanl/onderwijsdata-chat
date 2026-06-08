import asyncio
import io
import json
import os
import tempfile
import zipfile
from datetime import date

from dotenv import load_dotenv

load_dotenv()

import litellm
import chainlit as cl

import auth
import data_layer
from agent import generate_title, run
from chainlit.data import get_data_layer
from config import MODEL, WILLMA_API_KEY, get_available_models
from report import generate_pdf, generate_report
from resume import build_messages_from_thread, build_turns_from_thread
from tools import store
from tools.codegen import build_package

auth.setup()
data_layer.setup()

WELKOM = """Welkom! Ik kan je helpen met vragen over open Nederlandse onderwijsdata.

Ik heb toegang tot:
- **CBS** — statistieken over het Nederlandse onderwijs (68 datasets)
- **RIO** — register van onderwijsinstellingen en opleidingen (14 resources)
- **DUO** — 57 open datasets: prognoses, diplomering, instroom, adressen (onderwijsdata.duo.nl)

Stel een vraag, bijvoorbeeld:
- *Hoeveel MBO studenten waren er in 2023?*
- *Welke HBO-instellingen zijn er in Amsterdam?*
- *Wat is het verschil in uitstroom tussen mannen en vrouwen in het WO?*
- *Hoe ontwikkelt het MBO-studentenaantal zich richting 2040 per leerweg?*
"""

_FRIENDLY_ERRORS: list[tuple[type, str]] = [
    (litellm.AuthenticationError, "API key ontbreekt of is ongeldig. Controleer je `.env` bestand."),
    (litellm.NotFoundError, "Model niet gevonden. Controleer de `MODEL` instelling in `.env`."),
    (litellm.RateLimitError, "Te veel verzoeken naar de API. Wacht even en probeer opnieuw."),
    (litellm.APIConnectionError, "Kan de API niet bereiken. Controleer je internetverbinding."),
    (litellm.BadRequestError, "Het verzoek werd afgewezen door de API. Mogelijk een ongeldig model of parameter."),
]


def _friendly_error(exc: Exception) -> str:
    for exc_type, msg in _FRIENDLY_ERRORS:
        if isinstance(exc, exc_type):
            return f"❌ {msg}"
    return f"❌ Er is een fout opgetreden: {exc}"


async def _setup_modes() -> None:
    modus_mode = cl.Mode(
        id="modus",
        name="Modus",
        options=[
            cl.ModeOption(id="snel", name="Snel", description="Precies het gevraagde, niet meer", icon="zap", default=True),
            cl.ModeOption(id="verdiep", name="Verdiep", description="Doorvragen + volledige analyse", icon="microscope"),
        ],
    )

    modes = [modus_mode]

    available = get_available_models()
    if available:
        model_options = [
            cl.ModeOption(
                id=mid,
                name=name,
                description=desc,
                icon=icon,
                default=(mid == MODEL),
            )
            for mid, name, desc, icon in available
        ]
        if not any(o.default for o in model_options):
            model_options[0].default = True
        modes.insert(0, cl.Mode(id="model", name="Model", options=model_options))

    await cl.context.emitter.set_modes(modes)


async def _set_thread_title(question: str, answer: str, model: str | None = None) -> None:
    try:
        title = await generate_title(question, answer, model=model)
        layer = get_data_layer()
        if layer:
            await layer.update_thread(
                thread_id=cl.context.session.thread_id,
                name=title,
            )
        await cl.context.emitter.emit(
            "first_interaction",
            {"interaction": title, "thread_id": cl.context.session.thread_id},
        )
    except Exception:
        pass


@cl.set_starters
async def set_starters():
    return [
        cl.Starter(label="MBO studenten 2023", message="Hoeveel MBO studenten waren er in 2023?"),
        cl.Starter(label="HBO-instellingen Amsterdam", message="Welke HBO-instellingen zijn er in Amsterdam?"),
        cl.Starter(label="WO uitstroom man/vrouw", message="Wat is het verschil in uitstroom tussen mannen en vrouwen in het WO?"),
        cl.Starter(label="VMBO instroom trend", message="Toon de trend in VMBO instroom over de afgelopen 10 jaar."),
    ]


_UPLOAD_ROW_CAP = 10_000


async def _persist_uploads() -> None:
    import pandas as pd
    layer = get_data_layer()
    if not layer:
        return
    upload_keys = [k for k in store.list_keys() if k.startswith("upload:")]
    if not upload_keys:
        return
    serialized = {}
    for key in upload_keys:
        df = store.get(key)
        if not isinstance(df, pd.DataFrame):
            continue
        serialized[key] = df.head(_UPLOAD_ROW_CAP).to_csv(index=False)
    if serialized:
        await layer.update_thread(
            thread_id=cl.context.session.thread_id,
            metadata={"_uploads": serialized},
        )


@cl.on_chat_resume
async def on_chat_resume(thread: dict):
    import pandas as pd
    messages = build_messages_from_thread(thread)
    turns = build_turns_from_thread(thread)
    figures = [fig for turn in turns for fig in turn.get("figures", [])]
    cl.user_session.set("messages", messages)
    cl.user_session.set("turns", turns)
    cl.user_session.set("figures", figures)
    cl.user_session.set("turn_figures", [])

    raw_meta = thread.get("metadata") or {}
    if isinstance(raw_meta, str):
        try:
            raw_meta = json.loads(raw_meta)
        except Exception:
            raw_meta = {}
    for key, csv_str in (raw_meta.get("_uploads") or {}).items():
        try:
            df = pd.read_csv(io.StringIO(csv_str), dtype=str)
            store.put(key, df)
        except Exception:
            pass

    await _setup_modes()


@cl.on_chat_start
async def on_start():
    cl.user_session.set("messages", [])
    cl.user_session.set("figures", [])
    cl.user_session.set("turns", [])

    is_ollama = MODEL.startswith("ollama_chat/") or MODEL.startswith("ollama/")
    known_keys = ["ANTHROPIC_API_KEY", "OPENAI_API_KEY", "AZURE_API_KEY", "AZURE_AI_API_KEY", "GEMINI_API_KEY", "WILLMA_API_KEY"]
    if not is_ollama and not any(os.getenv(k) for k in known_keys):
        await cl.Message(content="⚠️ Geen API key gevonden. Stel een omgevingsvariabele in (bijv. `ANTHROPIC_API_KEY`) en herstart de app.").send()
        return

    await _setup_modes()


@cl.on_stop
async def on_stop():
    stop_event: asyncio.Event | None = cl.user_session.get("stop_event")
    if stop_event:
        stop_event.set()


async def _process_message(content: str, modus: str = "snel", model: str | None = None) -> None:
    cl.user_session.set("current_model", model)
    messages: list = cl.user_session.get("messages")
    messages.append({"role": "user", "content": content})

    stop_event = asyncio.Event()
    cl.user_session.set("stop_event", stop_event)
    cl.user_session.set("turn_figures", [])

    try:
        response_text = await run(messages, stop_event, model=model, modus=modus)
    except Exception as e:
        await cl.Message(content=_friendly_error(e)).send()
        return

    messages.append({"role": "assistant", "content": response_text})
    cl.user_session.set("messages", messages)

    turn_figures = cl.user_session.get("turn_figures", [])
    turn_tool_calls = cl.user_session.get("_last_turn_tool_calls", [])
    cl.user_session.set("_last_turn_tool_calls", [])
    turns: list = cl.user_session.get("turns", [])
    turns.append({"question": content, "answer": response_text, "figures": turn_figures, "tool_calls": turn_tool_calls})
    cl.user_session.set("turns", turns)

    if len(messages) == 2:
        asyncio.create_task(_set_thread_title(content, response_text, model=model))

    asyncio.create_task(_persist_uploads())


def _df_schema_json(df, data_key: str, name: str) -> str:
    schema = [
        {
            "kolom": col,
            "type": str(df[col].dtype),
            "uniek": int(df[col].nunique()),
            "voorbeelden": [str(v) for v in df[col].dropna().unique()[:3].tolist()],
        }
        for col in df.columns
    ]
    return json.dumps(
        {"data_key": data_key, "bestand": name, "totaal_rijen": len(df), "kolommen": schema},
        ensure_ascii=False, separators=(",", ":"),
    )


async def _read_file_content(el) -> str | None:
    path = el.path
    if not path:
        return None
    name = el.name or os.path.basename(path)
    lower = name.lower()

    try:
        import pandas as pd

        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            parts = []
            for sheet in sheet_names:
                df = pd.read_excel(path, sheet_name=sheet, dtype=str)
                if df.empty:
                    continue
                suffix = f":{sheet}" if len(sheet_names) > 1 else ""
                key = f"upload:{name}{suffix}"
                store.put(key, df)
                label = f"Sheet: {sheet} — " if len(sheet_names) > 1 else ""
                parts.append(f"{label}{_df_schema_json(df, key, name)}")
            return f"\n\n📎 **{name}**\n\n" + "\n\n".join(parts) if parts else None

        if lower.endswith(".csv"):
            df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
            if df.empty:
                return None
            key = f"upload:{name}"
            store.put(key, df)
            return f"\n\n📎 **{name}**\n\n{_df_schema_json(df, key, name)}"

    except Exception:
        return f"\n\n📎 **{name}** — kon niet worden gelezen."

    return None


@cl.on_message
async def on_message(message: cl.Message):
    modus = message.modes.get("modus", "snel") if message.modes else "snel"
    model = message.modes.get("model") if message.modes else None

    content = message.content
    if message.elements:
        for el in message.elements:
            file_text = await _read_file_content(el)
            if file_text:
                content += file_text

    await _process_message(content, modus=modus, model=model)


@cl.action_callback("followup")
async def on_followup(action: cl.Action):
    await cl.context.emitter.send_window_message({
        "type": "set_input",
        "value": action.payload["question"],
    })


@cl.action_callback("download_rapport")
async def on_download_rapport(action: cl.Action):
    turns = cl.user_session.get("turns", [])
    html = generate_report(turns)
    path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".html", delete=False, mode="w", encoding="utf-8") as f:
            f.write(html)
            path = f.name
        await cl.Message(
            content="Rapport klaar!",
            elements=[cl.File(name=f"rapport-{date.today()}.html", path=path, mime="text/html")],
        ).send()
    finally:
        if path:
            os.unlink(path)


@cl.action_callback("download_rapport_pdf")
async def on_download_rapport_pdf(action: cl.Action):
    turns = cl.user_session.get("turns", [])
    pdf_bytes = await asyncio.to_thread(generate_pdf, turns)
    path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(pdf_bytes)
            path = f.name
        await cl.Message(
            content="PDF rapport klaar!",
            elements=[cl.File(name=f"rapport-{date.today()}.pdf", path=path, mime="application/pdf")],
        ).send()
    finally:
        if path:
            os.unlink(path)


@cl.action_callback("download_python")
async def on_download_python(action: cl.Action):
    from agent import litellm_kwargs
    turns = cl.user_session.get("turns", [])
    thread_id = cl.context.session.thread_id or "chat"
    model = cl.user_session.get("current_model") or MODEL

    files = await build_package(turns, thread_id, model=model, extra_litellm_kwargs=litellm_kwargs(model))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for filename, content in files.items():
            zf.writestr(filename, content)
    zip_buf.seek(0)

    zip_name = f"analyse-{date.today()}-{thread_id[:8]}.zip"
    path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as f:
            f.write(zip_buf.read())
            path = f.name
        await cl.Message(
            content="Python pakket klaar!",
            elements=[cl.File(name=zip_name, path=path, mime="application/zip")],
        ).send()
    finally:
        if path:
            os.unlink(path)
