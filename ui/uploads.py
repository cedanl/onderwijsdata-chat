import json
import os

import chainlit as cl
from chainlit.data import get_data_layer

from tools import store

_UPLOAD_ROW_CAP = 10_000


def _df_schema_json(df, data_key: str, name: str) -> str:
    schema = [
        {
            "kolom": col,
            "type": str(df[col].dtype),
            "uniek": int(df[col].nunique()),
            "voorbeelden": [str(v) for v in df[col].dropna().unique()[:3].tolist()],
        }
        for col in df.columns
    ]
    return json.dumps(
        {"data_key": data_key, "bestand": name, "totaal_rijen": len(df), "kolommen": schema},
        ensure_ascii=False, separators=(",", ":"),
    )


def _upload_followup_actions(name: str, columns: list[str]) -> list[cl.Action]:
    return [
        cl.Action(name="followup", label=f"Samenvatting van {name}", payload={"question": f"Geef een statistische samenvatting van het bestand {name}."}),
        cl.Action(name="followup", label="Unieke waarden per kolom", payload={"question": f"Wat zijn de unieke waarden per kolom in {name}?"}),
        cl.Action(name="followup", label="Grafiek van verdeling", payload={"question": f"Maak een grafiek van de verdeling van {columns[0] if columns else 'de eerste kolom'} in {name}."}),
    ]


async def read_file_content(el) -> str | None:
    path = el.path
    if not path:
        return None
    name = el.name or os.path.basename(path)
    lower = name.lower()

    try:
        import pandas as pd

        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            llm_parts = []
            confirm_lines = [f"📎 **{name}** geladen"]
            all_columns: list[str] = []
            for sheet in sheet_names:
                df = pd.read_excel(path, sheet_name=sheet, dtype=str)
                if df.empty:
                    continue
                suffix = f":{sheet}" if len(sheet_names) > 1 else ""
                key = f"upload:{name}{suffix}"
                store.put(key, df)
                all_columns = list(df.columns)
                sheet_label = f" — sheet **{sheet}**" if len(sheet_names) > 1 else ""
                confirm_lines.append(f"{sheet_label}: {len(df):,} rijen, {len(df.columns)} kolommen")
                cols_preview = ", ".join(f"`{c}`" for c in df.columns[:8])
                if len(df.columns) > 8:
                    cols_preview += f" en {len(df.columns) - 8} meer"
                confirm_lines.append(f"Kolommen: {cols_preview}")
                llm_parts.append((f"Sheet: {sheet} — " if len(sheet_names) > 1 else "") + _df_schema_json(df, key, name))
            if not llm_parts:
                await cl.Message(content=f"📎 **{name}** — het bestand bevat geen data.").send()
                return None
            confirm_lines.append("\nJe kunt nu vragen stellen over dit bestand.")
            await cl.Message(
                content="\n".join(confirm_lines),
                actions=_upload_followup_actions(name, all_columns),
            ).send()
            return f"\n\n📎 **{name}**\n\n" + "\n\n".join(llm_parts)

        if lower.endswith(".csv"):
            df = pd.read_csv(path, encoding="utf-8-sig", dtype=str, sep=None, engine="python")
            if df.empty:
                await cl.Message(content=f"📎 **{name}** — het bestand bevat geen data.").send()
                return None
            key = f"upload:{name}"
            store.put(key, df)
            cols_preview = ", ".join(f"`{c}`" for c in df.columns[:8])
            if len(df.columns) > 8:
                cols_preview += f" en {len(df.columns) - 8} meer"
            await cl.Message(
                content=(
                    f"📎 **{name}** geladen — {len(df):,} rijen, {len(df.columns)} kolommen\n"
                    f"Kolommen: {cols_preview}\n\n"
                    "Je kunt nu vragen stellen over dit bestand."
                ),
                actions=_upload_followup_actions(name, list(df.columns)),
            ).send()
            return f"\n\n📎 **{name}**\n\n{_df_schema_json(df, key, name)}"

    except Exception as exc:
        await cl.Message(
            content=f"📎 **{name}** — kon niet worden ingelezen ({type(exc).__name__}). Controleer of het bestand niet beschadigd is en opgeslagen is als .csv of .xlsx."
        ).send()
        return None

    return None


async def persist_uploads() -> None:
    import pandas as pd
    layer = get_data_layer()
    if not layer:
        return
    upload_keys = [k for k in store.list_keys() if k.startswith("upload:")]
    if not upload_keys:
        return
    serialized = {}
    for key in upload_keys:
        df = store.get(key)
        if not isinstance(df, pd.DataFrame):
            continue
        serialized[key] = df.head(_UPLOAD_ROW_CAP).to_csv(index=False)
    if serialized:
        await layer.update_thread(
            thread_id=cl.context.session.thread_id,
            metadata={"_uploads": serialized},
        )
